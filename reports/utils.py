"""
Export utilities for generating PDF and Excel reports
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone


def generate_pdf_report(title, data, filename):
    """
    Generate a PDF report
    
    Args:
        title: Report title
        data: Dictionary containing report data
        filename: Output filename
    
    Returns:
        HttpResponse with PDF content
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4C1D95'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
        
        # Add date
        date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], alignment=TA_RIGHT, fontSize=10)
        generated_date = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
        elements.append(Paragraph(generated_date, date_style))
        elements.append(Spacer(1, 20))
        
        # Add summary if available
        if 'summary' in data:
            summary_data = []
            for key, value in data['summary'].items():
                summary_data.append([key, str(value)])
            
            summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F3F4F6')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.white)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
        
        # Add main data table if available
        if 'table' in data and data['table']:
            table = Table(data['table'], repeatRows=1)
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
            ]))
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF value
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(pdf)
        
        return response
        
    except ImportError:
        # Fallback if reportlab is not installed
        response = HttpResponse(content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{filename}.txt"'
        response.write(f"PDF export requires 'reportlab' package.\n\n")
        response.write(f"Install with: pip install reportlab\n\n")
        response.write(f"Report Title: {title}\n")
        response.write(f"Generated: {timezone.now()}\n")
        return response


def generate_excel_report(title, data, filename):
    """
    Generate an Excel report
    
    Args:
        title: Report title
        data: Dictionary containing report data
        filename: Output filename
    
    Returns:
        HttpResponse with Excel content
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Styling
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color="4C1D95")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Date
        ws.merge_cells('A2:E2')
        date_cell = ws['A2']
        date_cell.value = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal="right")
        
        current_row = 4
        
        # Add summary section
        if 'summary' in data:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            summary_title = ws[f'A{current_row}']
            summary_title.value = "Summary"
            summary_title.font = Font(bold=True, size=14)
            current_row += 1
            
            for key, value in data['summary'].items():
                ws[f'A{current_row}'] = key
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            
            current_row += 1
        
        # Add table data
        if 'table' in data and data['table']:
            table_data = data['table']
            
            # Headers
            for col_idx, header in enumerate(table_data[0], start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border_style
            
            current_row += 1
            
            # Data rows
            for row_data in table_data[1:]:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = cell_value
                    cell.border = border_style
                    
                    # Alternate row colors
                    if current_row % 2 == 0:
                        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                
                current_row += 1
            
            # Auto-adjust column widths
            for col_idx in range(1, len(table_data[0]) + 1):
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = 20
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        
        return response
        
    except ImportError:
        # Fallback if openpyxl is not installed
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        import csv
        writer = csv.writer(response)
        writer.writerow([title])
        writer.writerow([f"Generated: {timezone.now()}"])
        writer.writerow([])
        
        if 'summary' in data:
            writer.writerow(['Summary'])
            for key, value in data['summary'].items():
                writer.writerow([key, value])
            writer.writerow([])
        
        if 'table' in data:
            for row in data['table']:
                writer.writerow(row)
        
        return response
